import openpyxl
import time
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import ElementNotInteractableException, StaleElementReferenceException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

excel_file = 'Data Sample for Altro AI.xlsx'  
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook['REAL and Mocked up Data for POC'] 


link_column_index = None
for column in sheet.iter_cols(min_row=1, max_row=1):
    for cell in column:
        if cell.value and 'links' in str(cell.value).lower():
            link_column_index = column[0].column
            break

if link_column_index is None:
    print("Column 'links/sources' not found.")
else:
    urls = []
    for row in sheet.iter_rows(min_row=2):  
        link_cell = row[link_column_index - 1]  
        if link_cell.hyperlink:
            urls.append(link_cell.hyperlink.target)
    
    print("Extracted URLs:", urls)

    options = Options()
    options.add_argument("--headless")  
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

   
    csv_file = "scraped_data.csv"
    with open(csv_file, "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["URL", "Scraped Content", "Status"])  


        for url in urls:
            print(f"Navigating to {url}")
            driver.get(url)
            time.sleep(2)  
            status = "Failed"  

            try:
                   
                def find_about_menu():
                    try:
                    
                        about_menu = WebDriverWait(driver, 3).until(
                            EC.element_to_be_clickable((By.XPATH, "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'about')]"))
                        )
                        return about_menu
                    except Exception as e:
                        print("XPath search failed:", str(e))
                        about_variations = [
                            "About",   
                            "about",   
                            "ABOUT",   
                            "About Us",  
                            "about us",  
                            "ABOUT US",
                            "Who We Are"   
                        ]
                        for variation in about_variations:
                            try:
                                about_menu = WebDriverWait(driver, 1).until(
                                    EC.element_to_be_clickable((By.LINK_TEXT, variation))
                                )
                                return about_menu
                            except Exception as e:
                                print(f"Could not find {variation} link.")
                                continue
                        raise NoSuchElementException("Could not find 'About' or 'About Us' link")
                   

                about_menu = find_about_menu()
                about_link = about_menu.get_attribute("href")  

                
                before_click_url = driver.current_url
                about_menu.click()

                time.sleep(3)  

                after_click_url = driver.current_url

                if after_click_url == before_click_url or after_click_url == before_click_url + "#":
                    print("Clicking 'About' didn't navigate. Hovering instead.")
                
                    about_menu = find_about_menu()
                    
                    actions = ActionChains(driver)
                    actions.move_to_element(about_menu).perform()
                    time.sleep(2)  

                    first_sub_section = driver.find_element(By.XPATH, "//li[contains(@class, 'menu-item')]/a[contains(@href, 'about')]")
                    first_sub_link = first_sub_section.get_attribute("href")

                    print(f"First About Sub-Section: {first_sub_link}")

                    driver.get(first_sub_link)
                    time.sleep(3) 

                else:
                    print(f"Redirected to About page: {after_click_url}")

               
                about_content = driver.find_element(By.TAG_NAME, "body").text  
                print("Scraped Content:\n", about_content[:1000])  

                status = "Success"
                writer.writerow([url, about_content[:1000], status])

            except (StaleElementReferenceException, NoSuchElementException) as e:
                print(f"Error occurred while processing {url}: {str(e)}")
                status = "Failed"
                writer.writerow([url, "Error: " + str(e), status])

            except Exception as e:
                print(f"An unexpected error occurred while processing {url}: {str(e)}")
                status = "Failed"
                writer.writerow([url, "Error: " + str(e), status])


    driver.quit()

    print(f"Scraped data saved to {csv_file}")
